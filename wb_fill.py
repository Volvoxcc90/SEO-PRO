# wb_fill.py
from __future__ import annotations

import re
import json
import time
import random
from dataclasses import dataclass
from pathlib import Path
from typing import List, Optional, Callable, Set, Tuple

from openpyxl import load_workbook


# =========================================================
# Utils
# =========================================================
def _norm(s: str) -> str:
    s = (s or "").strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def _cap(s: str) -> str:
    s = (s or "").strip()
    return s[:1].upper() + s[1:] if s else s


def _safe_filename(name: str) -> str:
    name = re.sub(r'[\\/:*?"<>|]+', "_", name or "")
    name = re.sub(r"\s+", " ", name).strip()
    return name[:120] if name else "output"


def _first_sentences(text: str, n: int) -> str:
    parts = re.split(r"(?<=[.!?])\s+", text.strip())
    return " ".join(parts[:n]).strip()


def _jaccard(a: str, b: str) -> float:
    wa = set(re.findall(r"[a-zа-я0-9]+", a.lower()))
    wb = set(re.findall(r"[a-zа-я0-9]+", b.lower()))
    if not wa or not wb:
        return 0.0
    return len(wa & wb) / len(wa | wb)


def _join_ru(items: List[str]) -> str:
    items = [x.strip() for x in items if x.strip()]
    if not items:
        return ""
    if len(items) == 1:
        return items[0]
    if len(items) == 2:
        return f"{items[0]} и {items[1]}"
    return ", ".join(items[:-1]) + " и " + items[-1]


# =========================================================
# Params
# =========================================================
@dataclass
class FillParams:
    xlsx_path: str
    output_dir: str

    file_prefix: str
    start_index: int

    brand_lat: str
    brand_ru: str
    shape: str
    lenses: str
    collection: str

    holidays: str
    holiday_pos: str

    seo_level: str
    style: str

    wb_safe: bool
    wb_strict: bool

    brand_title_ratio: str

    rows_to_fill: int
    skip_first_rows: int
    batch_count: int

    fill_wb_fields: bool
    overwrite_wb_fields_if_not_empty: bool

    kiz: bool
    adult18: bool

    gender: str
    color: str
    composition: str

    progress_callback: Optional[Callable[[int], None]] = None


# =========================================================
# HUMAN SELLER ENGINE
# =========================================================

PERSONAS = {
    "casual": {
        "starters": [
            "Очки хорошо смотрятся на каждый день.",
            "Модель смотрится аккуратно и без лишнего перегруза.",
            "На лице очки выглядят легко и современно.",
            "Хороший вариант на повседневку и поездки.",
            "Очки смотрятся спокойно, но образ сразу становится интереснее.",
        ],
        "tone": "simple",
    },

    "fashion": {
        "starters": [
            "Очки сразу делают образ более стильным и заметным.",
            "Модель выглядит модно и хорошо вписывается в современные образы.",
            "Такие очки легко становятся акцентом в образе.",
            "Оправа смотрится эффектно и сразу притягивает внимание.",
        ],
        "tone": "fashion",
    },

    "mass": {
        "starters": [
            "Удобные очки на каждый день.",
            "Хороший вариант для города, прогулок и отдыха.",
            "Практичная модель, которую легко сочетать с одеждой.",
            "Очки подойдут и под обычный стиль, и под отпуск.",
        ],
        "tone": "mass",
    },

    "luxury": {
        "starters": [
            "Очки выглядят аккуратно и дорого.",
            "Модель смотрится сдержанно и стильно.",
            "Оправа выглядит выразительно, но без лишней яркости.",
            "Очки легко дополняют более собранные и дорогие образы.",
        ],
        "tone": "luxury",
    },
}


CHAOS_LINKS = [
    "Вообще",
    "Кстати",
    "При этом",
    "По ощущениям",
    "На деле",
    "Если носить каждый день",
    "В целом",
]


SEO_KEYS = [
    "очки солнцезащитные",
    "солнечные очки",
    "очки женские",
    "очки мужские",
    "очки унисекс",
    "очки для отпуска",
    "модные очки",
    "очки uv400",
    "брендовые очки",
]


# =========================================================
# TITLE ENGINE
# =========================================================
TITLE_STARTS = [
    "Стильные",
    "Модные",
    "Красивые",
    "Трендовые",
    "Удобные",
    "Летние",
    "Актуальные",
    "Эффектные",
    "Молодёжные",
]

TITLE_PRODUCTS = [
    "солнцезащитные очки",
    "солнечные очки",
]


def make_title(
    rnd: random.Random,
    p: FillParams,
    used_titles: Set[str],
) -> str:

    for _ in range(30):
        parts = []

        parts.append(rnd.choice(TITLE_STARTS))
        parts.append(rnd.choice(TITLE_PRODUCTS))

        include_brand = False

        ratio = (p.brand_title_ratio or "50/50").strip()

        if ratio == "100/0":
            include_brand = True
        elif ratio == "50/50":
            include_brand = rnd.random() < 0.5

        if include_brand and p.brand_ru:
            parts.append(p.brand_ru)

        if p.shape and rnd.random() < 0.45:
            parts.append(p.shape.lower())

        if p.lenses and rnd.random() < 0.45:
            parts.append(p.lenses)

        title = " ".join([x for x in parts if x]).strip()

        while len(title) > 60 and len(parts) > 2:
            parts.pop()
            title = " ".join(parts)

        sig = _norm(title)

        if sig not in used_titles:
            used_titles.add(sig)
            return title

    return title


# =========================================================
# HUMAN DESCRIPTION ENGINE
# =========================================================
def _pick_persona(style: str) -> str:
    style = (style or "").strip().lower()

    if style == "premium":
        return "luxury"

    if style == "mass":
        return "mass"

    if style == "social":
        return "fashion"

    return "casual"


def _seo_mix(rnd: random.Random, strict: bool) -> List[str]:
    keys = SEO_KEYS[:]

    if strict:
        keys = [k for k in keys if "инста" not in k and "tiktok" not in k]

    rnd.shuffle(keys)
    return keys[:3]


def _human_noise(rnd: random.Random) -> str:
    if rnd.random() < 0.45:
        return rnd.choice(CHAOS_LINKS) + ", "
    return ""


def make_description(
    rnd: random.Random,
    p: FillParams,
    used_descs: List[str],
    used_open2: Set[str],
    used_open3: Set[str],
) -> str:

    persona_name = _pick_persona(p.style)
    persona = PERSONAS[persona_name]

    for _ in range(40):

        blocks = []

        # =================================================
        # START
        # =================================================
        start = rnd.choice(persona["starters"])
        blocks.append(start)

        # =================================================
        # BRAND
        # =================================================
        if p.brand_lat and rnd.random() < 0.7:
            blocks.append(
                f"{_human_noise(rnd)}модель {p.brand_lat} легко сочетается с повседневной одеждой и не выглядит слишком тяжёлой."
            )

        # =================================================
        # SHAPE
        # =================================================
        if p.shape and rnd.random() < 0.8:
            shape_lines = [
                "Оправа подчёркивает лицо и выглядит аккуратно.",
                "Форма оправы смотрится современно и хорошо вписывается в разные стили.",
                "Оправа выглядит выразительно, но не перегружает образ.",
                "Такая форма оправы подходит под повседневный стиль и поездки.",
            ]
            blocks.append(_human_noise(rnd) + rnd.choice(shape_lines))

        # =================================================
        # LENSES
        # =================================================
        if p.lenses and rnd.random() < 0.8:

            lk = _norm(p.lenses)

            if "uv400" in lk:
                lens_line = "Линзы UV400 помогают чувствовать себя комфортно в солнечную погоду."

            elif "поляр" in lk:
                lens_line = "Поляризационные линзы уменьшают блики — особенно удобно за рулём и на улице."

            elif "хамелеон" in lk or "фотох" in lk:
                lens_line = "Фотохромные линзы подстраиваются под освещение, поэтому носить очки удобно в течение дня."

            else:
                lens_line = "Линзы делают использование очков более комфортным при ярком солнце."

            blocks.append(_human_noise(rnd) + lens_line)

        # =================================================
        # DAILY LIFE
        # =================================================
        if rnd.random() < 0.9:
            life_lines = [
                "Подойдут для прогулок, поездок, отдыха и обычного городского ритма.",
                "Очки удобно носить каждый день — и в городе, и на отдыхе.",
                "Модель подойдёт под повседневный стиль, отпуск и поездки.",
                "Очки хорошо смотрятся и с более спокойной одеждой, и с летними образами.",
            ]
            blocks.append(_human_noise(rnd) + rnd.choice(life_lines))

        # =================================================
        # GENDER
        # =================================================
        if p.gender:
            g = _norm(p.gender)

            if "уни" in g:
                blocks.append(
                    "Модель унисекс — хорошо подойдёт как девушкам, так и мужчинам."
                )

            elif "жен" in g:
                blocks.append(
                    "Очки хорошо дополняют женские образы и выглядят современно."
                )

            elif "муж" in g:
                blocks.append(
                    "Модель смотрится сдержанно и хорошо подходит под мужской стиль."
                )

        # =================================================
        # HOLIDAYS
        # =================================================
        if p.holidays:

            hs = [x.strip() for x in p.holidays.split("||") if x.strip()]

            if hs:
                holiday_line = (
                    f"Часто выбирают в подарок к {_join_ru(hs)}."
                )

                pos = (p.holiday_pos or "middle").lower()

                if pos == "start":
                    blocks.insert(1, holiday_line)

                elif pos == "end":
                    blocks.append(holiday_line)

                else:
                    blocks.insert(min(3, len(blocks)), holiday_line)

        # =================================================
        # SEO
        # =================================================
        keys = _seo_mix(rnd, p.wb_strict)

        if keys and rnd.random() < 0.9:
            seo_lines = [
                f"Подойдут тем, кто любит {keys[0]} без тяжёлой оправы.",
                f"Если нравятся {keys[0]}, эта модель хорошо впишется в гардероб.",
                f"Хороший вариант для тех, кто выбирает {keys[0]} на каждый день.",
            ]
            blocks.append(rnd.choice(seo_lines))

        # =================================================
        # RANDOM HUMAN CHAOS
        # =================================================
        rnd.shuffle(blocks[1:])

        # =================================================
        # DYNAMIC LENGTH
        # =================================================
        mode = rnd.choice(["short", "medium", "long"])

        if mode == "short":
            blocks = blocks[:4]

        elif mode == "medium":
            blocks = blocks[:6]

        # long = all

        # =================================================
        # BUILD
        # =================================================
        text = " ".join(
            _cap(x).rstrip(".") + "."
            for x in blocks if x.strip()
        )

        # =================================================
        # STRICT CLEAN
        # =================================================
        if p.wb_strict:

            text = re.sub(r"\bинста\b", "", text, flags=re.I)
            text = re.sub(r"\btiktok\b", "", text, flags=re.I)

            banned = [
                "идеально",
                "лучший",
                "самый лучший",
                "гарантия",
                "топ",
            ]

            for b in banned:
                text = re.sub(rf"\b{re.escape(b)}\b", "", text, flags=re.I)

        # =================================================
        # CLEAN
        # =================================================
        text = re.sub(r"\s{2,}", " ", text).strip()

        # =================================================
        # ANTI REPEATS
        # =================================================
        sig2 = _norm(_first_sentences(text, 2))
        sig3 = _norm(_first_sentences(text, 3))

        if sig2 in used_open2:
            continue

        if sig3 in used_open3:
            continue

        # similarity
        bad = False

        for old in used_descs[-30:]:

            if _jaccard(text, old) > 0.42:
                bad = True
                break

        if bad:
            continue

        used_open2.add(sig2)
        used_open3.add(sig3)
        used_descs.append(text)

        return text

    return text


# =========================================================
# EXCEL HELPERS
# =========================================================
def _detect_header_row(ws, scan: int = 30) -> int:

    for r in range(1, min(scan, ws.max_row) + 1):

        vals = []

        for c in range(1, min(80, ws.max_column) + 1):
            v = ws.cell(r, c).value
            if v is not None:
                vals.append(str(v))

        joined = _norm(" | ".join(vals))

        if "наимен" in joined and "описан" in joined:
            return r

    return 1


def _find_col(ws, header_row: int, needle: str):

    needle = _norm(needle)

    for c in range(1, ws.max_column + 1):

        v = ws.cell(header_row, c).value

        if v is None:
            continue

        if needle in _norm(str(v)):
            return c

    return None


def _set_cell(ws, r, c, value, overwrite: bool):

    cur = ws.cell(r, c).value

    if cur is None or str(cur).strip() == "":
        ws.cell(r, c).value = value
        return

    if overwrite:
        ws.cell(r, c).value = value


# =========================================================
# MAIN
# =========================================================
def fill_wb_template(params: FillParams):

    in_path = Path(params.xlsx_path)
    out_dir = Path(params.output_dir)

    out_dir.mkdir(parents=True, exist_ok=True)

    outputs = []

    used_titles = set()
    used_descs = []

    used_open2 = set()
    used_open3 = set()

    total_filled = 0

    for i in range(int(params.batch_count)):

        file_index = int(params.start_index) + i

        rnd = random.Random()

        rnd.seed(
            time.time_ns()
            ^ (file_index * 99991)
            ^ hash(params.brand_lat)
        )

        wb = load_workbook(in_path)
        ws = wb.active

        header_row = _detect_header_row(ws)

        col_name = _find_col(ws, header_row, "наименование")
        col_desc = _find_col(ws, header_row, "описание")

        if not col_name or not col_desc:
            raise ValueError("Не найдены колонки Наименование/Описание")

        # WB fields
        col_kiz = _find_col(ws, header_row, "киз")
        col_18 = _find_col(ws, header_row, "18+")
        col_gender = _find_col(ws, header_row, "пол")
        col_color = _find_col(ws, header_row, "цвет")
        col_comp = _find_col(ws, header_row, "состав")

        start_row = header_row + 1

        rows = [
            r for r in range(start_row, ws.max_row + 1)
            if r > int(params.skip_first_rows)
        ]

        rows = rows[:int(params.rows_to_fill)]

        for r in rows:

            title = make_title(
                rnd,
                params,
                used_titles,
            )

            desc = make_description(
                rnd,
                params,
                used_descs,
                used_open2,
                used_open3,
            )

            ws.cell(r, col_name).value = title
            ws.cell(r, col_desc).value = desc

            # WB fields
            if params.fill_wb_fields:

                ow = bool(params.overwrite_wb_fields_if_not_empty)

                if col_kiz:
                    _set_cell(ws, r, col_kiz, "да" if params.kiz else "нет", ow)

                if col_18:
                    _set_cell(ws, r, col_18, "да" if params.adult18 else "нет", ow)

                if col_gender and params.gender:
                    _set_cell(ws, r, col_gender, params.gender, ow)

                if col_color and params.color:
                    _set_cell(ws, r, col_color, params.color, ow)

                if col_comp and params.composition:
                    _set_cell(ws, r, col_comp, params.composition, ow)

        total_filled += len(rows)

        prefix = _safe_filename(params.file_prefix or in_path.stem)

        out_name = f"{prefix}_{file_index:04d}.xlsx"

        out_path = out_dir / out_name

        wb.save(out_path)

        outputs.append(str(out_path))

        if params.progress_callback:
            params.progress_callback(
                int((i + 1) * 100 / int(params.batch_count))
            )

    report = {
        "outputs": outputs,
        "rows_total_filled": total_filled,
    }

    return outputs, total_filled, json.dumps(report, ensure_ascii=False, indent=2)
