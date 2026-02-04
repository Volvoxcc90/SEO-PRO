import sys
from PyQt5.QtWidgets import (
    QApplication, QWidget, QLabel, QPushButton, QFileDialog, QLineEdit,
    QVBoxLayout, QHBoxLayout, QComboBox, QMessageBox, QProgressBar,
    QCheckBox, QDialog, QTextEdit, QDialogButtonBox, QTabWidget, QSplitter
)
from PyQt5.QtCore import QThread, pyqtSignal, Qt
from PyQt5.QtGui import QPalette, QColor, QLinearGradient, QBrush
from wb_fill import fill_wb_template
from utils import app_data_dir, setup_logging, load_list, auto_update_brand_map, add_to_list, logging
import openpyxl

class PreviewWorker(QThread):
    result = pyqtSignal(str)

    def __init__(self, args):
        super().__init__()
        self.args = args

    def run(self):
        try:
            wb = openpyxl.load_workbook(self.args['input_xlsx'], read_only=True)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            name_col = headers.index('Название') + 1 if 'Название' in headers else None
            desc_col = headers.index('Описание') + 1 if 'Описание' in headers else None
            if name_col or desc_col:
                model = ws.cell(2, name_col).value if name_col else ''
                current_desc = ws.cell(2, desc_col).value if desc_col else ''
                new_name = f"Пример: Солнцезащитные очки {self.args['brand']} {model} {self.args['collection']}"
                new_desc = f"Пример: {current_desc}\nSEO текст... (качество: {self.args['quality']}, длина: {self.args['length']})"
                preview = f"Новое название: {new_name}\nНовое описание: {new_desc}"
            else:
                preview = "Нет колонок для предпросмотра"
            self.result.emit(preview)
        except Exception as e:
            self.result.emit(str(e))

class Worker(QThread):
    progress = pyqtSignal(int)
    done = pyqtSignal(str, int, str)
    error = pyqtSignal(str)

    def __init__(self, args: dict):
        super().__init__()
        self.args = args

    def run(self):
        try:
            # Здесь расширь для новых args: quality, length и т.д.
            result = fill_wb_template(**self.args)
            out, count, report = result
            self.done.emit(out, count, report)
        except Exception as e:
            self.error.emit(str(e))

class ReportDialog(QDialog):
    def __init__(self, out: str, count: int, report: str, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Отчёт")
        self.resize(700, 500)
        layout = QVBoxLayout(self)
        info = QLabel(f"Файл создан: {out}\nСтрок: {count}")
        layout.addWidget(info)
        self.text = QTextEdit(report)
        self.text.setReadOnly(True)
        layout.addWidget(self.text)
        buttons = QDialogButtonBox(QDialogButtonBox.Close)
        buttons.rejected.connect(self.close)
        layout.addWidget(buttons)

class MainWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Sunglasses SEO PRO")
        self.resize(1000, 600)
        self.brands = load_list("brands.txt", ["Gucci", "Prada", "Miu Miu"])
        self.descriptions = load_list("descriptions.txt", ["Описания"])
        self.frames = load_list("frames.txt", ["Овальная"])
        auto_update_brand_map(self.brands)
        self._build()
        self._apply_dark_theme()

    def _apply_dark_theme(self):
        palette = QPalette()
        palette.setColor(QPalette.Window, QColor(25, 25, 35))
        palette.setColor(QPalette.WindowText, Qt.white)
        palette.setColor(QPalette.Base, QColor(35, 35, 45))
        palette.setColor(QPalette.Text, Qt.white)
        palette.setColor(QPalette.Button, QColor(50, 50, 60))
        palette.setColor(QPalette.ButtonText, Qt.white)
        app.setPalette(palette)
        grad = QLinearGradient(0, 0, 0, 1)
        grad.setColorAt(0, QColor(128, 0, 128))
        grad.setColorAt(1, QColor(75, 0, 75))
        self.run_button.setStyleSheet("background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #800080, stop:1 #4B004B); color: white; font-weight: bold; border-radius: 5px; padding: 10px;")

    def _build(self):
        main_layout = QVBoxLayout(self)
        title = QLabel("🕶 Sunglasses SEO PRO")
        title.setStyleSheet("font-size:22px; font-weight:700; color: #FFD700;")
        main_layout.addWidget(title)

        splitter = QSplitter(Qt.Horizontal)
        tabs = QTabWidget()
        basic_tab = QWidget()
        basic_layout = QVBoxLayout(basic_tab)
        theme_row = QHBoxLayout()
        self.theme = QComboBox()
        self.theme.addItems(["Midnight", "Dark", "Light"])
        theme_row.addWidget(QLabel("Тема"))
        theme_row.addWidget(self.theme)
        save_theme = QPushButton("Сохранить")
        theme_row.addWidget(save_theme)
        basic_layout.addLayout(theme_row)

        xlsx_row = QHBoxLayout()
        self.xlsx_lbl = QLabel("Бренд XLSX")
        load_xlsx = QPushButton("Загрузить XLSX")
        load_xlsx.clicked.connect(self.pick_file)
        xlsx_row.addWidget(self.xlsx_lbl)
        xlsx_row.addWidget(load_xlsx)
        basic_layout.addLayout(xlsx_row)

        brand_row = QHBoxLayout()
        self.brand = QComboBox()
        self.brand.addItems(self.brands)
        brand_row.addWidget(QLabel("Бренд"))
        brand_row.addWidget(self.brand)
        plus_brand = QPushButton("+")
        plus_brand.clicked.connect(lambda: self.add_item("brands.txt", self.brand))
        minus_brand = QPushButton("-")
        minus_brand.clicked.connect(lambda: self.remove_item("brands.txt", self.brand))
        brand_row.addWidget(plus_brand)
        brand_row.addWidget(minus_brand)
        basic_layout.addLayout(brand_row)

        desc_row = QHBoxLayout()
        self.desc = QComboBox()
        self.desc.addItems(self.descriptions)
        desc_row.addWidget(QLabel("Описания"))
        desc_row.addWidget(self.desc)
        plus_desc = QPushButton("+")
        plus_desc.clicked.connect(lambda: self.add_item("descriptions.txt", self.desc))
        minus_desc = QPushButton("-")
        minus_desc.clicked.connect(lambda: self.remove_item("descriptions.txt", self.desc))
        desc_row.addWidget(plus_desc)
        desc_row.addWidget(minus_desc)
        basic_layout.addLayout(desc_row)

        frame_row = QHBoxLayout()
        self.frame = QComboBox()
        self.frame.addItems(self.frames)
        frame_row.addWidget(QLabel("Форма оправы"))
        frame_row.addWidget(self.frame)
        plus_frame = QPushButton("+")
        plus_frame.clicked.connect(lambda: self.add_item("frames.txt", self.frame))
        minus_frame = QPushButton("-")
        minus_frame.clicked.connect(lambda: self.remove_item("frames.txt", self.frame))
        frame_row.addWidget(plus_frame)
        frame_row.addWidget(minus_frame)
        basic_layout.addLayout(frame_row)

        self.collection = QLineEdit("Весна-Лето 2026")
        basic_layout.addWidget(QLabel("Коллекция"))
        basic_layout.addWidget(self.collection)
        basic_layout.addStretch()

        seo_tab = QWidget()
        seo_layout = QVBoxLayout(seo_tab)
        self.quality = QComboBox()
        self.quality.addItems(["low", "medium", "premium"])
        seo_layout.addWidget(QLabel("Качество"))
        seo_layout.addWidget(self.quality)

        self.length = QComboBox()
        self.length.addItems(["short", "normal", "long"])
        seo_layout.addWidget(QLabel("Длина"))
        seo_layout.addWidget(self.length)

        self.style = QComboBox()
        self.style.addItems(["normal", "premium"])
        seo_layout.addWidget(QLabel("Стиль"))
        seo_layout.addWidget(self.style)
        seo_layout.addStretch()

        extra_tab = QWidget()
        extra_layout = QVBoxLayout(extra_tab)
        self.auto_ton = QComboBox()
        self.auto_ton.addItems(["Auto"])
        extra_layout.addWidget(QLabel("Auto-Ton"))
        extra_layout.addWidget(self.auto_ton)

        self.wb_safe = QComboBox()
        self.wb_safe.addItems(["WB Safe/Strict"])
        extra_layout.addWidget(QLabel("WB Safe"))
        extra_layout.addWidget(self.wb_safe)

        self.wb_strict = QComboBox()
        self.wb_strict.addItems(["WB Strict (яндекс анонсы/строн-паузы)"])
        extra_layout.addWidget(QLabel("WB Strict"))
        extra_layout.addWidget(self.wb_strict)
        extra_layout.addStretch()

        tabs.addTab(basic_tab, "Основные")
        tabs.addTab(seo_tab, "SEO-настройки")
        tabs.addTab(extra_tab, "Дополнительно")

        right = QWidget()
        right_layout = QVBoxLayout(right)
        preview_btn = QPushButton("🔍 Предпросмотр")
        preview_btn.clicked.connect(self.preview)
        right_layout.addWidget(preview_btn)
        self.preview_text = QTextEdit()
        self.preview_text.setReadOnly(True)
        right_layout.addWidget(self.preview_text)

        self.progress = QProgressBar()
        right_layout.addWidget(self.progress)

        splitter.addWidget(tabs)
        splitter.addWidget(right)
        splitter.setSizes([400, 600])
        main_layout.addWidget(splitter)

        footer = QHBoxLayout()
        self.run_button = QPushButton("🚀 ГЕНЕРИРОВАТЬ")
        self.run_button.clicked.connect(self.run)
        footer.addWidget(self.run_button)
        save_btn = QPushButton("Сохранить настройки")
        footer.addWidget(save_btn)
        main_layout.addLayout(footer)

    def add_item(self, filename, combo):
        val = combo.currentText().strip()
        if val:
            add_to_list(filename, val)
            items = load_list(filename, [])
            combo.clear()
            combo.addItems(items)
            combo.setCurrentText(val)

    def remove_item(self, filename, combo):
        val = combo.currentText().strip()
        if val:
            p = app_data_dir() / filename
            items = [x for x in load_list(filename, []) if x != val]
            p.write_text("\n".join(items), encoding="utf-8")
            combo.clear()
            combo.addItems(items)

    def pick_file(self):
        fp, _ = QFileDialog.getOpenFileName(self, "XLSX", "", "Excel (*.xlsx)")
        if fp:
            self.input_xlsx = fp
            self.xlsx_lbl.setText(fp)

    def preview(self):
        if not hasattr(self, "input_xlsx"):
            QMessageBox.warning(self, "Ошибка", "Загрузи XLSX")
            return
        args = {
            'input_xlsx': self.input_xlsx,
            'brand': self.brand.currentText(),
            'collection': self.collection.text(),
            'quality': self.quality.currentText(),
            'length': self.length.currentText()
            # Добавь другие
        }
        self.preview_worker = PreviewWorker(args)
        self.preview_worker.result.connect(self.preview_text.setPlainText)
        self.preview_worker.start()

    def run(self):
        if not hasattr(self, "input_xlsx"):
            QMessageBox.warning(self, "Ошибка", "Загрузи XLSX")
            return
        args = {
            'input_xlsx': self.input_xlsx,
            'brand': self.brand.currentText(),
            'collection': self.collection.text(),
            'quality': self.quality.currentText(),
            'length': self.length.currentText(),
            'style': self.style.currentText(),
            # Добавь auto_ton, wb_safe и т.д.
            'use_ai': False  # Если добавишь чекбокс
        }
        self.worker = Worker(args)
        self.worker.progress.connect(self.progress.setValue)
        self.worker.done.connect(lambda out, count, report: ReportDialog(out, count, report, self).exec_())
        self.worker.error.connect(lambda e: QMessageBox.critical(self, "Ошибка", e))
        self.worker.start()

if __name__ == "__main__":
    app = QApplication(sys.argv)
    w = MainWindow()
    w.show()
    sys.exit(app.exec_())
